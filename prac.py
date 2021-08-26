# from openpyxl import Workbook

# wb=Workbook()

# ws = wb.active()
# # ws1= wb.create_sheet("Mysheet")



# wb.save("test.xlsx")


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import datetime


wb = load_workbook(filename='sample.xlsx')
ws = wb.active
sheet_ranges=wb['Sheet1']

print(sheet_ranges['Q31'].value)

ws['Q32']=11111111


a3="SAMPLESAMPLE"
filename = datetime.datetime.now()
print(filename)
img = Image('logo.jpg')
ws.add_image(img, 'F3')
wb.save(a3+".xlsx")

# grab the active worksheet
# ws = wb.active

# ws.title="SHEET_12"
# # Data can be assigned directly to cells
# ws['A1'] = 42

# # Rows can also be appended
# ws.append([1, 2, 3])

# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# Save the file


# wb.save("estimate.xlsx")

